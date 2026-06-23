# -*- coding: utf-8 -*-
"""
process_ecommerce_sales.py

Lê sales_data.csv, limpa e valida os dados, calcula as agregações
de negócio e gera um relatório Excel com 4 abas visíveis:

    Resumo Executivo  -> KPIs principais
    Dados             -> transações já com cabeçalhos em português
    Análise           -> tabelas agregadas (mês, produto, categoria, cidade, frete, desconto)
    Dashboard         -> 5 gráficos baseados nas tabelas da aba Análise

Uma 5ª aba (Base dos Gráficos) é criada apenas como fonte de dados
para os gráficos e fica oculta no arquivo final — Excel não permite
que um gráfico leia direto de uma tabela com título mesclado, então
os dados "limpos" (sem título) ficam replicados ali.

Saídas: sales_processed.csv, sales_report_FINAL.xlsx
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

FONTE = 'Calibri'
AZUL = '2E5395'
VERDE = '70AD47'
CINZA_CLARO = 'E7E6E6'
BORDA_FINA = Border(*([Side(style='thin', color='D9D9D9')] * 4))


def carregar_e_limpar(caminho_csv='sales_data.csv'):
    """Lê o CSV bruto, remove inválidos/duplicados e adiciona a coluna auxiliar de mês."""
    df = pd.read_csv(caminho_csv, encoding='utf-8-sig')
    df['Data'] = pd.to_datetime(df['Data'])
    df = df.dropna(subset=['ID_Pedido', 'Data', 'Produto', 'Valor_Final'])
    df = df[df['Quantidade'] > 0]
    df = df[df['Valor_Final'] > 0]
    df = df.drop_duplicates(subset='ID_Pedido')
    df['Mes'] = df['Data'].dt.strftime('%Y-%m')  # usada só para o groupby por mês
    return df


def traduzir_para_negocio(df):
    """Renomeia as colunas técnicas para cabeçalhos em português de negócio (aba Dados)."""
    return df.rename(columns={
        'ID_Pedido': 'Pedido',
        'Data': 'Data',
        'Produto': 'Produto',
        'Categoria': 'Categoria',
        'Preco_Unitario': 'Preço Unitário (R$)',
        'Quantidade': 'Quantidade',
        'Cidade': 'Cidade',
        'Tipo_Cliente': 'Tipo de Cliente',
        'Desconto_Pct': '% Desconto',
        'Frete': 'Frete',
        'Valor_Bruto': 'Valor Bruto (R$)',
        'Valor_Final': 'Valor Final (R$)',
    })[['Pedido', 'Data', 'Produto', 'Categoria', 'Preço Unitário (R$)', 'Quantidade',
        'Cidade', 'Tipo de Cliente', '% Desconto', 'Frete', 'Valor Bruto (R$)', 'Valor Final (R$)']]


def calcular_agregacoes(df):
    """
    Calcula todas as tabelas agregadas de negócio.
    Os valores são arredondados aqui, no cálculo — não só na formatação de
    exibição — para que o valor "cru" da célula já seja o valor correto
    (evita casos como 3436.06091 aparecer ao copiar a célula).
    """
    vendas_mes = df.groupby('Mes', as_index=False)['Valor_Final'].agg(['sum', 'count', 'mean'])
    vendas_mes.columns = ['Mês', 'Vendas Totais (R$)', 'Qtd. Pedidos', 'Ticket Médio (R$)']
    vendas_mes[['Vendas Totais (R$)', 'Ticket Médio (R$)']] = vendas_mes[['Vendas Totais (R$)', 'Ticket Médio (R$)']].round(2)
    vendas_mes = vendas_mes.sort_values('Mês')

    top_produtos = df.groupby('Produto', as_index=False)['Valor_Final'].agg(['sum', 'count', 'mean'])
    top_produtos.columns = ['Produto', 'Receita (R$)', 'Qtd. Pedidos', 'Ticket Médio (R$)']
    top_produtos[['Receita (R$)', 'Ticket Médio (R$)']] = top_produtos[['Receita (R$)', 'Ticket Médio (R$)']].round(2)
    top_produtos = top_produtos.sort_values('Receita (R$)', ascending=False).head(5)

    vendas_categoria = df.groupby('Categoria', as_index=False)['Valor_Final'].agg(['sum', 'count'])
    vendas_categoria.columns = ['Categoria', 'Receita (R$)', 'Qtd. Pedidos']
    vendas_categoria['Receita (R$)'] = vendas_categoria['Receita (R$)'].round(2)
    vendas_categoria = vendas_categoria.sort_values('Receita (R$)', ascending=False)

    vendas_cidade = df.groupby('Cidade', as_index=False)['Valor_Final'].agg(['sum', 'count', 'mean'])
    vendas_cidade.columns = ['Cidade', 'Receita (R$)', 'Qtd. Pedidos', 'Ticket Médio (R$)']
    vendas_cidade[['Receita (R$)', 'Ticket Médio (R$)']] = vendas_cidade[['Receita (R$)', 'Ticket Médio (R$)']].round(2)
    vendas_cidade = vendas_cidade.sort_values('Receita (R$)', ascending=False)

    frete = df.groupby('Frete', as_index=False)['Valor_Final'].agg(['sum', 'count', 'mean'])
    frete.columns = ['Frete', 'Receita (R$)', 'Qtd. Pedidos', 'Ticket Médio (R$)']
    frete[['Receita (R$)', 'Ticket Médio (R$)']] = frete[['Receita (R$)', 'Ticket Médio (R$)']].round(2)

    desconto = df.groupby('Desconto_Pct', as_index=False)['Valor_Final'].agg(['sum', 'count', 'mean'])
    desconto.columns = ['% Desconto', 'Receita (R$)', 'Qtd. Pedidos', 'Ticket Médio (R$)']
    desconto[['Receita (R$)', 'Ticket Médio (R$)']] = desconto[['Receita (R$)', 'Ticket Médio (R$)']].round(2)
    desconto = desconto.sort_values('% Desconto')

    kpis = {
        'total_vendas': round(float(df['Valor_Final'].sum()), 2),
        'total_pedidos': len(df),
        'ticket_medio': round(float(df['Valor_Final'].mean()), 2),
        'clientes_novos': int((df['Tipo_Cliente'] == 'Novo').sum()),
        'clientes_recorrentes': int((df['Tipo_Cliente'] == 'Recorrente').sum()),
    }
    kpis['pct_novos'] = round(kpis['clientes_novos'] / kpis['total_pedidos'] * 100, 1)
    kpis['pct_recorrentes'] = round(kpis['clientes_recorrentes'] / kpis['total_pedidos'] * 100, 1)

    return {
        'vendas_mes': vendas_mes,
        'top_produtos': top_produtos,
        'vendas_categoria': vendas_categoria,
        'vendas_cidade': vendas_cidade,
        'frete': frete,
        'desconto': desconto,
        'kpis': kpis,
    }


def aplica_estilo_cabecalho(celula, texto, cor_fundo=AZUL, tamanho=11, cor_fonte='FFFFFF'):
    """Formata uma célula como cabeçalho: fundo colorido, texto branco em negrito."""
    celula.value = texto
    celula.font = Font(name=FONTE, bold=True, color=cor_fonte, size=tamanho)
    celula.fill = PatternFill('solid', start_color=cor_fundo)
    celula.alignment = Alignment(horizontal='center', vertical='center')


def escreve_tabela(ws, titulo, tabela, linha_inicial, colunas_moeda=None, colunas_pct=None):
    """
    Escreve uma tabela (DataFrame) na planilha `ws`, a partir de `linha_inicial`,
    com um título mesclado, cabeçalho verde e linhas com borda fina.
    Aplica formato de moeda (R$) ou percentual nas colunas indicadas.
    Retorna a próxima linha livre, para a próxima tabela poder começar abaixo.
    """
    colunas_moeda = colunas_moeda or []
    colunas_pct = colunas_pct or []
    n_colunas = len(tabela.columns)

    aplica_estilo_cabecalho(ws.cell(row=linha_inicial, column=1), titulo, cor_fundo=AZUL, tamanho=12)
    ws.merge_cells(start_row=linha_inicial, start_column=1, end_row=linha_inicial, end_column=n_colunas)

    for col_idx, nome_coluna in enumerate(tabela.columns, 1):
        aplica_estilo_cabecalho(ws.cell(row=linha_inicial + 1, column=col_idx), nome_coluna, cor_fundo=VERDE, tamanho=10)

    for i, linha in enumerate(tabela.itertuples(index=False), start=2):
        for col_idx, valor in enumerate(linha, 1):
            celula = ws.cell(row=linha_inicial + i, column=col_idx, value=valor)
            celula.font = Font(name=FONTE, size=10)
            celula.border = BORDA_FINA
            nome_coluna = tabela.columns[col_idx - 1]
            if nome_coluna in colunas_moeda:
                celula.number_format = 'R$ #,##0.00'
                celula.alignment = Alignment(horizontal='right')
            elif nome_coluna in colunas_pct:
                celula.number_format = '0"%"'
                celula.alignment = Alignment(horizontal='center')
            elif 'Qtd' in nome_coluna:
                celula.number_format = '#,##0'
                celula.alignment = Alignment(horizontal='right')
            else:
                celula.alignment = Alignment(horizontal='left')

    return linha_inicial + len(tabela) + 3  # +3 = título + cabeçalho + 1 linha de espaço


def copia_tabela_para_base(ws_base, tabela, linha_atual):
    """Copia um DataFrame para ws_base a partir de linha_atual, devolve (inicio, fim, proxima_linha)."""
    inicio = linha_atual
    for linha_dados in dataframe_to_rows(tabela, index=False, header=True):
        for c_idx, valor in enumerate(linha_dados, 1):
            ws_base.cell(row=linha_atual, column=c_idx, value=valor)
        linha_atual += 1
    return inicio, linha_atual - 1, linha_atual + 1


def montar_resumo_executivo(wb, kpis):
    """Cria a aba 'Resumo Executivo' com os KPIs principais."""
    ws = wb.create_sheet('Resumo Executivo')
    ws.sheet_properties.tabColor = AZUL

    ws['A1'] = 'ANÁLISE DE VENDAS E-COMMERCE — BRASIL'
    ws['A1'].font = Font(name=FONTE, size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color=AZUL)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 28

    linhas_kpi = [
        ('Receita Total', kpis['total_vendas'], 'R$ #,##0.00', AZUL),
        ('Total de Pedidos', kpis['total_pedidos'], '#,##0', VERDE),
        ('Ticket Médio', kpis['ticket_medio'], 'R$ #,##0.00', '4472C4'),
        ('Clientes Novos', f"{kpis['clientes_novos']} ({kpis['pct_novos']:.1f}%)", None, 'FFC000'),
        ('Clientes Recorrentes', f"{kpis['clientes_recorrentes']} ({kpis['pct_recorrentes']:.1f}%)", None, 'C55A11'),
    ]
    linha = 3
    for rotulo, valor, formato, cor in linhas_kpi:
        ws.cell(row=linha, column=1, value=rotulo).font = Font(name=FONTE, bold=True, color='FFFFFF', size=11)
        ws.cell(row=linha, column=1).fill = PatternFill('solid', start_color=cor)
        celula_valor = ws.cell(row=linha, column=2, value=valor)
        celula_valor.font = Font(name=FONTE, bold=True, size=13)
        celula_valor.fill = PatternFill('solid', start_color=CINZA_CLARO)
        celula_valor.alignment = Alignment(horizontal='right')
        if formato:
            celula_valor.number_format = formato
        ws.row_dimensions[linha].height = 24
        linha += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 24
    ws.cell(row=linha + 1, column=1,
            value='Período analisado: últimos 6 meses · Dataset sintético gerado para fins de portfólio.'
            ).font = Font(name=FONTE, italic=True, size=9, color='808080')
    return ws


def montar_dados(wb, df_negocio):
    """Cria a aba 'Dados' com as 1.000 transações em formato de negócio."""
    ws = wb.create_sheet('Dados')
    ws.sheet_properties.tabColor = VERDE

    for r_idx, linha_dados in enumerate(dataframe_to_rows(df_negocio, index=False, header=True), 1):
        for c_idx, valor in enumerate(linha_dados, 1):
            celula = ws.cell(row=r_idx, column=c_idx, value=valor)
            if r_idx == 1:
                aplica_estilo_cabecalho(celula, valor, cor_fundo=AZUL, tamanho=9)
                celula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                celula.font = Font(name=FONTE, size=9)
                celula.border = BORDA_FINA
                nome_coluna = df_negocio.columns[c_idx - 1]
                if nome_coluna == 'Data':
                    celula.number_format = 'dd/mm/yyyy'
                    celula.alignment = Alignment(horizontal='center')
                elif 'R$' in nome_coluna:
                    celula.number_format = 'R$ #,##0.00'
                    celula.alignment = Alignment(horizontal='right')
                elif nome_coluna == '% Desconto':
                    celula.number_format = '0"%"'
                    celula.alignment = Alignment(horizontal='center')
                else:
                    celula.alignment = Alignment(horizontal='left')

    ws.freeze_panes = 'A2'
    larguras = [12, 12, 11, 13, 16, 11, 15, 15, 11, 9, 15, 15]
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = largura
    return ws


def montar_analise(wb, agregacoes):
    """Cria a aba 'Análise' com as 6 tabelas agregadas de negócio."""
    ws = wb.create_sheet('Análise')
    ws.sheet_properties.tabColor = 'FFC000'

    linha = 1
    linha = escreve_tabela(ws, 'Vendas por Mês', agregacoes['vendas_mes'], linha,
                            colunas_moeda=['Vendas Totais (R$)', 'Ticket Médio (R$)'])
    linha = escreve_tabela(ws, 'Top 5 Produtos por Receita', agregacoes['top_produtos'], linha,
                            colunas_moeda=['Receita (R$)', 'Ticket Médio (R$)'])
    linha = escreve_tabela(ws, 'Vendas por Categoria', agregacoes['vendas_categoria'], linha,
                            colunas_moeda=['Receita (R$)'])
    linha = escreve_tabela(ws, 'Vendas por Cidade', agregacoes['vendas_cidade'], linha,
                            colunas_moeda=['Receita (R$)', 'Ticket Médio (R$)'])
    linha = escreve_tabela(ws, 'Frete Grátis vs. Pago', agregacoes['frete'], linha,
                            colunas_moeda=['Receita (R$)', 'Ticket Médio (R$)'])
    linha = escreve_tabela(ws, 'Desconto Aplicado', agregacoes['desconto'], linha,
                            colunas_moeda=['Receita (R$)', 'Ticket Médio (R$)'], colunas_pct=['% Desconto'])

    for col, largura in zip(['A', 'B', 'C', 'D'], [22, 18, 15, 18]):
        ws.column_dimensions[col].width = largura
    return ws


def montar_dashboard(wb, agregacoes):
    """Cria a aba auxiliar oculta 'Base dos Gráficos' e a aba 'Dashboard' com os 5 gráficos."""
    ws_base = wb.create_sheet('Base dos Gráficos')
    ws_base.sheet_properties.tabColor = '808080'

    r = 1
    mes_ini, mes_fim, r = copia_tabela_para_base(ws_base, agregacoes['vendas_mes'], r)
    prod_ini, prod_fim, r = copia_tabela_para_base(ws_base, agregacoes['top_produtos'], r)
    cat_ini, cat_fim, r = copia_tabela_para_base(ws_base, agregacoes['vendas_categoria'], r)
    cid_ini, cid_fim, r = copia_tabela_para_base(ws_base, agregacoes['vendas_cidade'], r)
    frete_ini, frete_fim, r = copia_tabela_para_base(ws_base, agregacoes['frete'], r)

    ws_dash = wb.create_sheet('Dashboard')
    ws_dash.sheet_properties.tabColor = 'C55A11'

    ws_dash['A1'] = 'DASHBOARD — VENDAS E-COMMERCE'
    ws_dash['A1'].font = Font(name=FONTE, size=15, bold=True, color='FFFFFF')
    ws_dash['A1'].fill = PatternFill('solid', start_color=AZUL)
    ws_dash.merge_cells('A1:P1')
    ws_dash.row_dimensions[1].height = 26

    # Gráfico 1: receita mensal (linha)
    grafico1 = LineChart()
    grafico1.title = 'Receita Mensal'
    grafico1.style = 12
    grafico1.y_axis.title = 'Receita (R$)'
    grafico1.x_axis.title = 'Mês'
    grafico1.height, grafico1.width = 9, 16
    dados1 = Reference(ws_base, min_col=2, min_row=mes_ini, max_row=mes_fim)
    categorias1 = Reference(ws_base, min_col=1, min_row=mes_ini + 1, max_row=mes_fim)
    grafico1.add_data(dados1, titles_from_data=True)
    grafico1.set_categories(categorias1)
    grafico1.legend = None  # série única; o eixo X já identifica os meses, legenda é redundante
    ws_dash.add_chart(grafico1, 'A3')

    # Gráfico 2: top 5 produtos (barras verticais)
    grafico2 = BarChart()
    grafico2.type = 'col'
    grafico2.title = 'Top 5 Produtos por Receita'
    grafico2.style = 10
    grafico2.y_axis.title = 'Receita (R$)'
    grafico2.height, grafico2.width = 9, 16
    dados2 = Reference(ws_base, min_col=2, min_row=prod_ini, max_row=prod_fim)
    categorias2 = Reference(ws_base, min_col=1, min_row=prod_ini + 1, max_row=prod_fim)
    grafico2.add_data(dados2, titles_from_data=True)
    grafico2.set_categories(categorias2)
    grafico2.legend = None  # série única; o eixo X já identifica os produtos
    ws_dash.add_chart(grafico2, 'I3')

    # Gráfico 3: receita por categoria (pizza)
    grafico3 = PieChart()
    grafico3.title = 'Receita por Categoria'
    grafico3.style = 26
    grafico3.height, grafico3.width = 9, 16
    dados3 = Reference(ws_base, min_col=2, min_row=cat_ini, max_row=cat_fim)
    categorias3 = Reference(ws_base, min_col=1, min_row=cat_ini + 1, max_row=cat_fim)
    grafico3.add_data(dados3, titles_from_data=True)
    grafico3.set_categories(categorias3)
    # Pizza sem legenda visível é ilegível (só cores, sem saber qual é qual).
    # Mostramos só categoria + percentual nas fatias — não valor nem nome da série,
    # que deixam o rótulo poluído e sobreposto.
    from openpyxl.chart.label import DataLabelList
    grafico3.dataLabels = DataLabelList()
    grafico3.dataLabels.showCatName = True
    grafico3.dataLabels.showPercent = True
    grafico3.dataLabels.showVal = False
    grafico3.dataLabels.showSerName = False
    grafico3.dataLabels.showLegendKey = False
    grafico3.legend = None  # info já está nas próprias fatias, legenda separada seria redundante
    ws_dash.add_chart(grafico3, 'A21')

    # Gráfico 4: receita por cidade (barras horizontais)
    grafico4 = BarChart()
    grafico4.type = 'bar'
    grafico4.title = 'Receita por Cidade'
    grafico4.style = 11
    grafico4.x_axis.title = 'Receita (R$)'
    grafico4.height, grafico4.width = 9, 16
    dados4 = Reference(ws_base, min_col=2, min_row=cid_ini, max_row=cid_fim)
    categorias4 = Reference(ws_base, min_col=1, min_row=cid_ini + 1, max_row=cid_fim)
    grafico4.add_data(dados4, titles_from_data=True)
    grafico4.set_categories(categorias4)
    grafico4.legend = None  # série única; o eixo já identifica as cidades
    ws_dash.add_chart(grafico4, 'I21')

    # Gráfico 5: ticket médio, frete grátis vs. pago (barras verticais)
    grafico5 = BarChart()
    grafico5.type = 'col'
    grafico5.title = 'Ticket Médio: Frete Grátis vs. Pago'
    grafico5.style = 14
    grafico5.y_axis.title = 'Ticket Médio (R$)'
    grafico5.height, grafico5.width = 9, 16
    dados5 = Reference(ws_base, min_col=4, min_row=frete_ini, max_row=frete_fim)
    categorias5 = Reference(ws_base, min_col=1, min_row=frete_ini + 1, max_row=frete_fim)
    grafico5.add_data(dados5, titles_from_data=True)
    grafico5.set_categories(categorias5)
    grafico5.legend = None  # série única; o eixo já identifica Grátis/Pago
    ws_dash.add_chart(grafico5, 'A39')

    # A aba auxiliar só deve ficar oculta DEPOIS de os gráficos já terem sido
    # criados — se for ocultada antes, alguns visualizadores de Excel não
    # conseguem renderizar os gráficos corretamente.
    ws_base.sheet_state = 'hidden'
    return ws_dash


def main():
    df = carregar_e_limpar('sales_data.csv')
    df.to_csv('sales_processed.csv', index=False, encoding='utf-8-sig')
    print(f'OK: {len(df)} registros limpos -> sales_processed.csv')

    df_negocio = traduzir_para_negocio(df)
    agregacoes = calcular_agregacoes(df)
    print(f"OK: agregações calculadas. Receita total = R$ {agregacoes['kpis']['total_vendas']:,.2f}")

    wb = Workbook()
    wb.remove(wb.active)  # remove a aba "Sheet" padrão criada pelo openpyxl

    montar_resumo_executivo(wb, agregacoes['kpis'])
    montar_dados(wb, df_negocio)
    montar_analise(wb, agregacoes)
    montar_dashboard(wb, agregacoes)

    wb.save('sales_report_FINAL.xlsx')
    print('OK: sales_report_FINAL.xlsx gerado (Resumo Executivo + Dados + Análise + Dashboard)')


if __name__ == '__main__':
    main()
